"""
Skapar Testprotokoll_Strategiportfoljen_v310.xlsx
Täcker alla funktioner i v3.10 — används av Martin för manuell verifiering.
"""
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

NAVY      = "1E3A5F"; ACCENT = "0EA5E9"; SUITE_BG = "2563EB"
ROW_ALT   = "F0F7FF"; ROW_WHITE = "FFFFFF"; GREEN_BG = "DCFCE7"
RED_BG    = "FEE2E2"; YELLOW_BG = "FEF9C3"; GRAY_BG = "F3F4F6"
BORDER_C  = "E5E7EB"

def fill(c): return PatternFill("solid", fgColor=c)
def bdr():
    s = Side(style="thin", color=BORDER_C)
    return Border(top=s, bottom=s, left=s, right=s)
def hfont(sz=11, bold=True, col="FFFFFF"): return Font(name="Calibri", size=sz, bold=bold, color=col)
def bfont(sz=10, bold=False, col="111827"): return Font(name="Calibri", size=sz, bold=bold, color=col)
def cell(ws, r, c, v="", bg=None, fnt=None, aln=None):
    x = ws.cell(row=r, column=c, value=v)
    if bg:  x.fill = fill(bg)
    if fnt: x.font = fnt
    if aln: x.alignment = aln
    x.border = bdr()
    return x

C = Alignment(horizontal="center", vertical="center")
L = Alignment(horizontal="left", vertical="center", wrap_text=True)
TL= Alignment(horizontal="left", vertical="top", wrap_text=True)

wb = Workbook()

# ── Sammanfattnings-flik ──────────────────────────────────────────────
ws0 = wb.active; ws0.title = "Sammanfattning"
ws0.column_dimensions["A"].width = 6
ws0.column_dimensions["B"].width = 36
ws0.column_dimensions["C"].width = 14
ws0.column_dimensions["D"].width = 14
ws0.column_dimensions["E"].width = 16

ws0.merge_cells("B2:E2")
cell(ws0,2,2,"Testprotokoll — Strategiportföljen v3.10", NAVY, hfont(14), C)
ws0.row_dimensions[2].height = 30

ws0.merge_cells("B3:E3")
cell(ws0,3,2,"Fyll i resultat löpande — OK / Fel / Ej testat", ACCENT, hfont(10,False), C)

for c,h in zip([2,3,4,5],["Testsvit","Antal testfall","Godkända","Underkända"]):
    cell(ws0,5,c,h,NAVY,hfont(),C)
ws0.row_dimensions[5].height = 22

sviter = [
    ("T1 — Importflöde & ordningsguide",  8),
    ("T2 — Dashboard & portföljdiagram",  9),
    ("T3 — Kassa & Avstämning",           8),
    ("T4 — Innehav & Signaler",           7),
    ("T5 — Excel-backup & Restore",      10),
    ("T6 — Kontoregister & namnbyte",     5),
    ("T7 — Mobil / iPad",                 6),
    ("T8 — Veckorutin (helflöde)",        7),
    ("T9 — Felhantering",                 6),
]
totalt = sum(n for _,n in sviter)
for i,(namn,antal) in enumerate(sviter):
    r = 6+i; bg = ROW_ALT if i%2==0 else ROW_WHITE
    cell(ws0,r,2,namn,bg,bfont(),L)
    cell(ws0,r,3,antal,bg,bfont(),C)
    cell(ws0,r,4,"",bg,bfont(),C)
    cell(ws0,r,5,"",bg,bfont(),C)
r = 6+len(sviter)
cell(ws0,r,2,"TOTALT",NAVY,hfont(),L)
cell(ws0,r,3,totalt,NAVY,hfont(),C)

ws0.row_dimensions[7+len(sviter)].height = 8
cell(ws0,9+len(sviter),2,"Testat av:",GRAY_BG,bfont(bold=True),L)
cell(ws0,9+len(sviter),3,"",ROW_WHITE,bfont(),L); ws0.merge_cells(f"C{9+len(sviter)}:E{9+len(sviter)}")
cell(ws0,10+len(sviter),2,"Datum:",GRAY_BG,bfont(bold=True),L)
cell(ws0,10+len(sviter),3,"",ROW_WHITE,bfont(),L); ws0.merge_cells(f"C{10+len(sviter)}:E{10+len(sviter)}")
cell(ws0,11+len(sviter),2,"Version testad:",GRAY_BG,bfont(bold=True),L)
cell(ws0,11+len(sviter),3,"v3.10",ROW_WHITE,bfont(),L)

# ── Hjälpfunktion: skapa testsvit-flik ───────────────────────────────
def svit(title, code, rubrik, beskrivning, testfall):
    """
    testfall: lista av (id, beskrivning, steg, förväntat_resultat)
    """
    ws = wb.create_sheet(title)
    COLS = {"A":6,"B":14,"C":46,"D":36,"E":14,"F":14}
    for col,w in COLS.items(): ws.column_dimensions[col].width = w

    # Rubrikrad
    ws.merge_cells("B2:F2")
    cell(ws,2,2,f"{code} — {rubrik}",SUITE_BG,hfont(12),C)
    ws.row_dimensions[2].height = 28
    ws.merge_cells("B3:F3")
    cell(ws,3,2,beskrivning,"F0F7FF",bfont(9,False,"374151"),TL)
    ws.row_dimensions[3].height = 40

    # Kolumnhuvud
    for c,h in zip([2,3,4,5,6],["Test-ID","Testbeskrivning","Teststeg","Förväntat resultat","Resultat"]):
        cell(ws,5,c,h,NAVY,hfont(10),C)
    ws.row_dimensions[5].height = 20

    # Dropdown för Resultat
    dv = DataValidation(type="list", formula1='"OK,Fel,Ej testat"', allow_blank=True)
    ws.add_data_validation(dv)

    for i,(tid, tbesk, tsteg, tfört) in enumerate(testfall):
        r = 6+i; bg = ROW_ALT if i%2==0 else ROW_WHITE
        ws.row_dimensions[r].height = 48
        cell(ws,r,2,tid,bg,bfont(9,True,"1E3A5F"),C)
        cell(ws,r,3,tbesk,bg,bfont(9),TL)
        cell(ws,r,4,tsteg,bg,bfont(9,"374151"),TL)
        cell(ws,r,5,tfört,bg,bfont(9,"15803d"),TL)
        res = cell(ws,r,6,"Ej testat",YELLOW_BG,bfont(9,True,"374151"),C)
        dv.add(res)

    return ws

# ════════════════════════════════════════════════════════════════════
#  T1 — Importflöde & ordningsguide
# ════════════════════════════════════════════════════════════════════
svit("T1 – Importflöde", "T1", "Importflöde & ordningsguide",
     "Testar att filer importeras korrekt, att ordningsguiden visar rätt status och att varningar visas vid fel ordning.",
[
    ("T1.1","Ordningsguide visas på Importera-fliken",
     "1. Gå till Importera-fliken\n2. Titta längst upp",
     "Kortet '📋 Importordning — veckorutin' visas med 4 steg, alla ⬜ om inget importerats"),
    ("T1.2","Transaktionsfil importeras (steg 1)",
     "1. Tryck Transaktioner\n2. Välj transaktionsfil från Avanza\n3. Kolla ordningsguiden",
     "Steg 1 visar ✅ och senaste importdatum. Importloggen visar antal transaktioner."),
    ("T1.3","Positionsfil importeras (steg 2)",
     "1. Tryck Positioner\n2. Välj positionsfil\n3. Kolla logg och guide",
     "Steg 2 visar ✅. Logg visar '💡 Tips: Gå till Avstämning...'"),
    ("T1.4","Varning vid inköpskurs utan positioner",
     "1. Rensa appen (Backup → Rensa)\n2. Försök importera inköpskursfil UTAN att ha positioner",
     "❌-felmeddelande: 'Inga innehav — importera en positionsfil (steg 2) först'"),
    ("T1.5","Inköpskurs efter positioner (steg 3)",
     "1. Importera positionsfil\n2. Importera inköpskursfil\n3. Kolla steg 3 i guiden",
     "Steg 3 visar ✅. GAV-historik uppdateras för matchande innehav."),
    ("T1.6","Flera positionsfiler bygger historik",
     "1. Importera 3 positionsfiler med olika datum\n2. Gå till Dashboard → Allt",
     "Värdeutvecklingsdiagrammet visar 3+ datapunkter, en per importerad fil."),
    ("T1.7","Sålda innehav tas bort",
     "1. Importera äldre positionsfil med fler innehav\n2. Importera nyare utan ett innehav",
     "Innehav som saknas i senaste filen tas bort och loggas: '📤 Borttagna: ...'"),
    ("T1.8","Dubbel import hanteras",
     "1. Importera samma transaktionsfil två gånger",
     "Inga dubbletter — antal transaktioner ökar inte vid andra importen"),
])

# ════════════════════════════════════════════════════════════════════
#  T2 — Dashboard & portföljdiagram
# ════════════════════════════════════════════════════════════════════
svit("T2 – Dashboard", "T2", "Dashboard & portföljdiagram",
     "Testar det interaktiva diagrammet, nyckeltalskorten, period-val och kategoriutveckling.",
[
    ("T2.1","Huvud-diagram visas med default 'i År'",
     "1. Gå till Dashboard\n2. Titta längst upp",
     "Diagram visas direkt. Aktiv period-knapp = 'i År'. Blå linje för portföljvärde."),
    ("T2.2","Periodknappar byter diagramperiod",
     "1. Tryck '30D'\n2. Tryck 'Allt'",
     "Diagrammet uppdateras till rätt period. Aktiv knapp markeras. Titeln uppdateras."),
    ("T2.3","Kryssrutor lägger till/tar bort serier",
     "1. Kryssa i 'Nettoinsatt kapital'\n2. Kryssa i 'Nettoresultat'",
     "Grå streckad linje (Nettoinsatt) och grön linje (Nettoresultat) tillkommer i diagrammet"),
    ("T2.4","Kategorilinjer aktiveras",
     "1. Kryssa i en kategori (t.ex. ⚓ Ankaret)",
     "En linje i kategorifärgen tillkommer om historik med katVärden finns"),
    ("T2.5","Stapeldiagram fungerar",
     "1. Tryck '📊 Stapel'-knappen",
     "Diagrammet byter till staplar. Ikonen/knappen markeras som aktiv."),
    ("T2.6","Nyckeltalskort visar korrekt data",
     "1. Kontrollera att 7 kort visas\n2. Kontrollera att vart och ett har ikon, värde och förklaringsrad",
     "💼 Portföljvärde, 📈 Avkastning, 💳 Nettoinsatt, 📊 Innehav, 🏦 Kassa, 🌱 Utdelningar, ✅ Realiserat"),
    ("T2.7","Global period ändrar nyckeltal",
     "1. Välj '30D' i globala period-strippen\n2. Kolla avkastningskortet",
     "Avkastnings-kortet visar 'Avkastning · 30D' och ett 30-dagars värde"),
    ("T2.8","Signalband visas vid röda signaler",
     "1. Skapa ett innehav med kurs under MA200 i kat. 3–6 och tvådagars aktiv",
     "Rött band visas ovanför nyckeltal med '🔴 1 säljsignal'. Klick → Signaler-fliken."),
    ("T2.9","Tomt diagram hanteras",
     "1. Rensa historik (inget importerat)\n2. Gå till Dashboard",
     "Diagrammet visar text 'Importera positionsfiler för att se historik'"),
])

# ════════════════════════════════════════════════════════════════════
#  T3 — Kassa & Avstämning
# ════════════════════════════════════════════════════════════════════
svit("T3 – Kassa och Avstämning", "T3", "Kassa & Avstämning",
     "Testar att kassan hämtas automatiskt från positionsfilen och att diff-wizarden stämmer av mot Avanza.",
[
    ("T3.1","Tillgängligt för köp hämtas från positionsfil",
     "1. Importera positionsfil\n2. Gå till Kassa-fliken",
     "Tabellen visar 5 ISK/KF-konton (ej sparkonto) med 📄-ikon och belopp från positionsfilen"),
    ("T3.2","Sparkonto visas INTE i Kassa-tabellen",
     "1. Gå till Kassa-fliken\n2. Titta i 'Tillgängligt för köp'-tabellen",
     "'Avanza sparande Martin' visas INTE i tabellen — det visas i Avstämning istället"),
    ("T3.3","Alla 6 konton i Avstämning",
     "1. Gå till Avstämning-fliken\n2. Titta på Saldon per konto",
     "Alla 6 konton visas en gång vardera. Sparkontot visar kontonummer som subrad."),
    ("T3.4","Diff-wizard steg 2 — Avanza-navigering",
     "1. Klicka 'Starta →' i Avstämning\n2. Gå till steg 2",
     "Blå inforuta: 'Min ekonomi → Sparande → Totalt sparande'. Två inmatningsfält."),
    ("T3.5","Diff-wizard steg 3 — liten diff",
     "1. Ange siffror från Avanza i steg 2\n2. Tryck 'Beräkna diff →'\n3. Diff < 200 kr",
     "Grönt summakort: '✅ Allt stämmer'. Grön beslutsgraf utan varningar."),
    ("T3.6","Diff-wizard steg 3 — stor diff utan positionsfil",
     "1. Rensa appen\n2. Ange Avanza-totaler utan att ha importerat positioner",
     "Röd åtgärd med prioritet 1️⃣: 'Ingen positionsfil importerad — importera...'"),
    ("T3.7","Manuell override i Kassa",
     "1. Välj ett konto i Kassa-formuläret\n2. Ange ett belopp och spara",
     "Kontot visar ✅-ikon. Värdet används istället för positionsfilens 📄-värde."),
    ("T3.8","Summan i Avstämning matchar Avanza",
     "1. Importera positionsfil\n2. Jämför Avstämnings-totalen med Avanza 'Sparande'",
     "Diff under 200 kr (FX-avrundning). Kontrollrutan visar ✅."),
])

# ════════════════════════════════════════════════════════════════════
#  T4 — Innehav & Signaler
# ════════════════════════════════════════════════════════════════════
svit("T4 – Innehav & Signaler", "T4", "Innehav & Signaler",
     "Testar innehavstabellen, MA200-signaler, nödutgångar och kategoritilldelning.",
[
    ("T4.1","Innehavstabellen visar rätt kolumner",
     "1. Importera positionsfil\n2. Gå till Innehav",
     "Tabell: Namn, Kategori, Antal, GAV (SEK), Kurs, MA200, Signal, Andel. Sorterbar per kolumn."),
    ("T4.2","Grön MA200-signal fungerar",
     "1. Öppna ett innehav\n2. Sätt MA200 = 80 kr, kurs = 100 kr (>5% över)",
     "Signalcell visar 🟢. Dashboard: inga röda signaler för det innehavet."),
    ("T4.3","Röd signal kräver tvådagarssregel (kat. 3–6)",
     "1. Kat. 3–innehavet: MA200 = 100 kr, kurs = 90 kr\n2. Importera samma fil dagen efter",
     "Dag 1: 🟡 (gul). Dag 2+: 🔴 (röd) och signalbandet visar '🔴 1 säljsignal'"),
    ("T4.4","Nödutgång hård visas vid GAV−10%",
     "1. Innehav kat. 4: GAV = 100 kr, kurs = 89 kr (under 90% av GAV)",
     "Signal: 🛑 Nödutgång hård. Signalbandet visar '🛑 1 nödutgång'."),
    ("T4.5","Ticker-sökning fungerar",
     "1. Gå till Innehav → 📡 Tickers & MA200\n2. Klicka 🔍 för ett innehav\n3. Sök 'NVIDIA'",
     "Lista med börsalternativ visas. Val uppdaterar ticker-fältet."),
    ("T4.6","MA200-hämtning via Alpha Vantage",
     "1. Ange Alpha Vantage API-nyckel\n2. Klicka 'Uppdatera alla MA200'",
     "MA200-värden hämtas och uppdateras. Inget ⚠-valutavarning för SEK-innehav."),
    ("T4.7","FX-motor: utländsk aktie visar bolagsvinst/FX",
     "1. Gå till ett USD-innehav\n2. Klicka 💰 (utdelningar/detaljer)",
     "Bolagsvinst och FX-vinst visas separat. Historisk FX = GAV(SEK)/GAV(USD)."),
])

# ════════════════════════════════════════════════════════════════════
#  T5 — Excel-backup & Restore
# ════════════════════════════════════════════════════════════════════
svit("T5 – Excel Backup & Restore", "T5", "Excel-backup & Restore",
     "Testar att all data exporteras korrekt och kan återställas fullständigt — inklusive historik och kategoriutveckling.",
[
    ("T5.1","Export skapar fil med 7 ark",
     "1. Gå till Importera → Backup\n2. Tryck 'Ladda ner Excel-backup'",
     "Excel-fil laddas ned. Öppna: ska ha 7 ark: Innehav, Transaktioner, Beslutslogg, Sammanfattning, Historik, Kassa, Kategorier"),
    ("T5.2","Historik-arket innehåller totalVärde (ej tomt)",
     "1. Öppna Excel-filen\n2. Gå till arket 'Historik'",
     "Kolumn B (Värde) innehåller faktiska tal — INTE tomma celler. Kolumn D (KatVärden) innehåller JSON."),
    ("T5.3","Kassa-arket bevarar kontoStartsaldo",
     "1. Ange manuellt saldo för ett konto i Kassa\n2. Exportera\n3. Öppna Excel",
     "Arket 'Kassa' innehåller sektion 'KontoStartsaldo' med kontot och beloppet"),
    ("T5.4","Kategorier-arket bevarar anpassade kategorier",
     "1. Redigera en kategori (ändra färg/mål)\n2. Exportera\n3. Öppna Excel",
     "Arket 'Kategorier' visar de anpassade värdena"),
    ("T5.5","Full restore: data återställs",
     "1. Exportera backup\n2. Rensa ALLT (localStorage)\n3. Importera backup-filen",
     "Alla innehav, transaktioner, logg och historik återställs. Antal stämmer med Sammanfattning-arket."),
    ("T5.6","Historik-diagrammet fungerar efter restore",
     "1. Utför T5.5 (full restore)\n2. Gå till Dashboard → Allt",
     "Värdeutvecklingsdiagrammet visar historiklinjen — INTE tomt. Portföljvärdet stämmer."),
    ("T5.7","Kategoriutveckling fungerar efter restore",
     "1. Utför T5.5\n2. Gå till Kategorier → Jämför kategorier",
     "Kategoriutvecklingsdiagrammet visar linjer (kräver katVärden i historik)"),
    ("T5.8","Kategorier återställs från backup",
     "1. Utför T5.5\n2. Gå till Kategorier → Hantera",
     "Kategorier med anpassade färger/mål återställs korrekt"),
    ("T5.9","Restore rensar positionsKassa korrekt",
     "1. Utför T5.5\n2. Gå till Kassa",
     "Kassa-tabellen visar manuella saldo (✅) om sådana sparats. Positionsfil-värden (📄) kräver ny import."),
    ("T5.10","Backup på ny enhet",
     "1. Exportera på enhet A\n2. Importera på enhet B (annan webbläsare)\n3. Kontrollera Dashboard",
     "Portföljvärde, antal innehav och historik stämmer. Positionsfil kan behöva importeras på nytt."),
])

# ════════════════════════════════════════════════════════════════════
#  T6 — Kontoregister & namnbyte
# ════════════════════════════════════════════════════════════════════
svit("T6 – Kontoregister", "T6", "Kontoregister & namnbyte i Avanza",
     "Testar att appen automatiskt hanterar när ett konto döps om i Avanza.",
[
    ("T6.1","Kontoregistret byggs vid positionsimport",
     "1. Importera en positionsfil\n2. Kontrollera i Konsolen: localStorage.getItem('portfölj_konto_register')",
     "JSON-objekt med kontonummer som nycklar och {senastNamn, senastSett, historia:[]} per konto"),
    ("T6.2","Korrekt namn visas i Avstämning",
     "1. Importera positionsfil\n2. Gå till Avstämning → Saldon per konto",
     "Kontonamnen matchar Avanzas namn (t.ex. '1. Sverige nov 2025')"),
    ("T6.3","Namnbyte detekteras vid nästa import",
     "1. Simulera namnbyte: redigera en rad i positionsfil-CSV\n2. Importera den ändrade filen",
     "Kontoregistret uppdateras: senastNamn = nytt namn, historia = [{gammal, till: datum}]"),
    ("T6.4","Historikindikator visas i Avstämning",
     "1. Utför T6.3\n2. Gå till Avstämning",
     "Kontot visar '🔁 1 tidigare namn' med tooltip: 'Gammalt namn (t.o.m. DATUM)'"),
    ("T6.5","avanzaKontoInfo hittar konto via register",
     "1. Utför T6.3 (namnbyte)\n2. Gå till Kassa och Innehav",
     "Konto identifieras korrekt via register — rätt typ (ISK/KF/SPAR) och kontonummer"),
])

# ════════════════════════════════════════════════════════════════════
#  T7 — Mobil / iPad
# ════════════════════════════════════════════════════════════════════
svit("T7 – Mobil iPad", "T7", "Mobilanvändning på iPad",
     "Testar att alla funktioner fungerar på iPad i Safari och att touch-ytor är tillräckliga.",
[
    ("T7.1","Filimport fungerar i Safari på iPad",
     "1. Öppna appen i Safari på iPad\n2. Gå till Importera\n3. Tryck på 'Positioner'-rutan",
     "Filväljar-dialog öppnas i iOS Filer-appen. Fil kan väljas och importeras."),
    ("T7.2","Flervalsimport fungerar",
     "1. Tryck på 'Positioner'\n2. Välj flera filer (tryck länge eller använd välj-knapp)",
     "Alla valda filer importeras. Loggen visar en rad per fil."),
    ("T7.3","Dashboard-diagrammet är touchvänligt",
     "1. Öppna Dashboard på iPad\n2. Tryck på diagrammet",
     "Tooltip visas vid tryck. Diagram är responsivt och fyller skärmbredden."),
    ("T7.4","Period-knappar kan tryckas på iPad",
     "1. Tryck på period-knapp i huvud-diagrammet på iPad",
     "Knappen aktiveras och diagrammet uppdateras. Knapp-yta ≥ 44px (touch-standard)."),
    ("T7.5","Mobilkort visas i Innehav",
     "1. Gå till Innehav på iPad (smal vy)\n2. Sök efter ett innehav",
     "Innehav visas som mobil-kort (ej tabell) med alla relevanta fält"),
    ("T7.6","Bottom-navigering fungerar",
     "1. Tryck på flikarna i bottom-navigeringen\n2. Byt flik",
     "Rätt sektion visas. Aktiv flik markeras. Smooth-scroll till toppen."),
])

# ════════════════════════════════════════════════════════════════════
#  T8 — Veckorutin (helflöde)
# ════════════════════════════════════════════════════════════════════
svit("T8 – Veckorutin", "T8", "Veckorutin — komplett helflöde",
     "Simulerar hela veckorutinen från export i Avanza till loggpost.",
[
    ("T8.1","Steg 1: Exportera filer från Avanza",
     "1. Öppna avanza.se i Safari\n2. Min ekonomi → Transaktioner → Exportera\n3. Min ekonomi → Innehav → Exportera",
     "Två filer sparas i Filer-appen: transaktioner_DATUM.csv och positioner_DATUM.csv"),
    ("T8.2","Steg 2: Importera transaktioner (steg 1 i guiden)",
     "1. Gå till Importera → Transaktioner\n2. Välj transaktionsfilen",
     "✅ grön i ordningsguiden steg 1. Logg: 'X nya transaktioner importerade'"),
    ("T8.3","Steg 2: Importera positioner (steg 2 i guiden)",
     "1. Gå till Importera → Positioner\n2. Välj positionsfilen",
     "✅ grön i steg 2. Logg: kassa per konto + historikpunkt tillagd + Avstämnings-tips"),
    ("T8.4","Steg 3: Stäm av mot Avanza",
     "1. Gå till Avstämning\n2. Starta diff-kontrollen\n3. Ange Avanza-totaler",
     "Diff < 1% (acceptabel). Grön status i steg 3."),
    ("T8.5","Steg 3: Kontrollera signalband",
     "1. Gå till Dashboard\n2. Läs signalbandet om det visas",
     "Inga röda signaler ELLER: röd signal identifierad och kan hanteras via Signaler-fliken"),
    ("T8.6","Steg 3: Kontrollera kategoribalans",
     "1. Gå till Kategorier\n2. Kolla viktbalansen",
     "Inga kategorier utanför målintervall ELLER: ombalanseringsbehov visas tydligt med kr-belopp"),
    ("T8.7","Steg 4: Skriv beslutslogg",
     "1. Gå till Beslutslogg\n2. Fyll i vecka, datum och portföljvärde\n3. Skriv tankar och åtgärder\n4. Spara",
     "Loggpost sparas och visas i listan. Exportera till Excel fungerar."),
])

# ════════════════════════════════════════════════════════════════════
#  T9 — Felhantering
# ════════════════════════════════════════════════════════════════════
svit("T9 – Felhantering", "T9", "Felhantering & felmeddelanden",
     "Testar att appen hanterar fel gracefully med tydliga felmeddelanden.",
[
    ("T9.1","Fel filtyp ger tydligt felmeddelande",
     "1. Tryck på Positioner-import\n2. Välj en transaktionsfil (fel typ)",
     "Felmeddelande: '❌ Fel filformat! Det här verkar inte vara en positionsfil.'"),
    ("T9.2","Inköpskurs utan positioner blockeras",
     "1. Rensa appen\n2. Försök importera inköpskursfil",
     "'❌ Inga innehav — importera en positionsfil (steg 2) först'"),
    ("T9.3","Okänt konto filtreras vid import",
     "1. Lägg till rad med okänt kontonummer i positionsfil-CSV\n2. Importera",
     "Loggen: 'X värdepapper exkluderade av filter'. Okänt konto importeras inte."),
    ("T9.4","Excel-backup med fel format ger felmeddelande",
     "1. Välj en icke-backup Excel-fil som backup",
     "Felmeddelande om saknade ark (Innehav, Transaktioner etc.)"),
    ("T9.5","Tomt portfölj hanteras utan krasch",
     "1. Rensa all data\n2. Gå igenom alla flikar",
     "Alla flikar visar tomma tillstånd med förklarande text. Inga JavaScript-fel i konsolen."),
    ("T9.6","API-fel vid MA200-hämtning",
     "1. Ange ogiltig API-nyckel\n2. Tryck 'Uppdatera alla MA200'",
     "Felmeddelande per innehav. Inga innehav uppdateras. Appen kraschar inte."),
])

# ── Spara ──────────────────────────────────────────────────────────
dst = r"C:\Users\hejma\Projekt_Claude\strategiportfoljen\Testprotokoll_Strategiportfoljen_v310.xlsx"
wb.save(dst)
print(f"Skapad: {dst}")
print(f"Totalt: {sum(n for _,n in sviter)} testfall i 9 sviter")
