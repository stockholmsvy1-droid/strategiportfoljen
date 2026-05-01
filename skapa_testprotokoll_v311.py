"""
Skapar Testprotokoll_Strategiportfoljen_v311.xlsx
Täcker alla funktioner i v3.11 — används av Martin för manuell verifiering.
"""
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

NAVY      = "1E3A5F"; ACCENT = "0EA5E9"; SUITE_BG = "2563EB"
ROW_ALT   = "F0F7FF"; ROW_WHITE = "FFFFFF"; GREEN_BG = "DCFCE7"
RED_BG    = "FEE2E2"; YELLOW_BG = "FEF9C3"; GRAY_BG = "F3F4F6"
BORDER_C  = "E5E7EB"

def fill(c): return PatternFill("solid", fgColor=c)
def bdr():
    s = Side(style="thin", color=BORDER_C)
    return Border(top=s, bottom=s, left=s, right=s)
def hfont(sz=11, bold=True, col="FFFFFF"): return Font(name="Calibri", size=sz, bold=bold, color=col)
def bfont(sz=10, bold=False, col="111827"): return Font(name="Calibri", size=sz, bold=bold, color=col)
def cell(ws, r, c, v="", bg=None, fnt=None, aln=None):
    x = ws.cell(row=r, column=c, value=v)
    if bg:  x.fill = fill(bg)
    if fnt: x.font = fnt
    if aln: x.alignment = aln
    x.border = bdr()
    return x

C = Alignment(horizontal="center", vertical="center")
L = Alignment(horizontal="left", vertical="center", wrap_text=True)
TL= Alignment(horizontal="left", vertical="top", wrap_text=True)

wb = Workbook()

ws0 = wb.active; ws0.title = "Sammanfattning"
ws0.column_dimensions["A"].width = 6
ws0.column_dimensions["B"].width = 36
ws0.column_dimensions["C"].width = 14
ws0.column_dimensions["D"].width = 14
ws0.column_dimensions["E"].width = 16

ws0.merge_cells("B2:E2")
cell(ws0,2,2,"Testprotokoll — Strategiportföljen v3.11", NAVY, hfont(14), C)
ws0.row_dimensions[2].height = 30

ws0.merge_cells("B3:E3")
cell(ws0,3,2,"Fyll i resultat löpande — OK / Fel / Ej testat", ACCENT, hfont(10,False), C)

for c,h in zip([2,3,4,5],["Testsvit","Antal testfall","Godkända","Underkända"]):
    cell(ws0,5,c,h,NAVY,hfont(),C)
ws0.row_dimensions[5].height = 22

sviter = [
    ("T1 — Importflöde & ordningsguide",    8),
    ("T2 — Dashboard & portföljdiagram",    9),
    ("T3 — Kassa & Avstämning",             9),
    ("T4 — Innehav & Signaler",             7),
    ("T5 — Excel-backup & Restore",        10),
    ("T6 — Kontoregister & namnbyte",       5),
    ("T7 — Mobil / iPad",                   6),
    ("T8 — Veckorutin (helflöde)",          7),
    ("T9 — Felhantering",                   6),
    ("T10 — Inställningar-sektionen",      10),
]
totalt = sum(n for _,n in sviter)
for i,(namn,antal) in enumerate(sviter):
    r = 6+i; bg = ROW_ALT if i%2==0 else ROW_WHITE
    cell(ws0,r,2,namn,bg,bfont(),L)
    cell(ws0,r,3,antal,bg,bfont(),C)
    cell(ws0,r,4,"",bg,bfont(),C)
    cell(ws0,r,5,"",bg,bfont(),C)
r = 6+len(sviter)
cell(ws0,r,2,"TOTALT",NAVY,hfont(),L)
cell(ws0,r,3,totalt,NAVY,hfont(),C)

ws0.row_dimensions[7+len(sviter)].height = 8
cell(ws0,9+len(sviter),2,"Testat av:",GRAY_BG,bfont(bold=True),L)
cell(ws0,9+len(sviter),3,"",ROW_WHITE,bfont(),L); ws0.merge_cells(f"C{9+len(sviter)}:E{9+len(sviter)}")
cell(ws0,10+len(sviter),2,"Datum:",GRAY_BG,bfont(bold=True),L)
cell(ws0,10+len(sviter),3,"",ROW_WHITE,bfont(),L); ws0.merge_cells(f"C{10+len(sviter)}:E{10+len(sviter)}")
cell(ws0,11+len(sviter),2,"Version testad:",GRAY_BG,bfont(bold=True),L)
cell(ws0,11+len(sviter),3,"v3.11",ROW_WHITE,bfont(),L)

def svit(title, code, rubrik, beskrivning, testfall):
    ws = wb.create_sheet(title)
    COLS = {"A":6,"B":14,"C":46,"D":36,"E":14,"F":14}
    for col,w in COLS.items(): ws.column_dimensions[col].width = w
    ws.merge_cells("B2:F2")
    cell(ws,2,2,f"{code} — {rubrik}",SUITE_BG,hfont(12),C)
    ws.row_dimensions[2].height = 28
    ws.merge_cells("B3:F3")
    cell(ws,3,2,beskrivning,"F0F7FF",bfont(9,False,"374151"),TL)
    ws.row_dimensions[3].height = 40
    for c,h in zip([2,3,4,5,6],["Test-ID","Testbeskrivning","Teststeg","Förväntat resultat","Resultat"]):
        cell(ws,5,c,h,NAVY,hfont(10),C)
    ws.row_dimensions[5].height = 20
    dv = DataValidation(type="list", formula1='"OK,Fel,Ej testat"', allow_blank=True)
    ws.add_data_validation(dv)
    for i,(tid, tbesk, tsteg, tfört) in enumerate(testfall):
        r = 6+i; bg = ROW_ALT if i%2==0 else ROW_WHITE
        ws.row_dimensions[r].height = 48
        cell(ws,r,2,tid,bg,bfont(9,True,"1E3A5F"),C)
        cell(ws,r,3,tbesk,bg,bfont(9),TL)
        cell(ws,r,4,tsteg,bg,bfont(9,"374151"),TL)
        cell(ws,r,5,tfört,bg,bfont(9,"15803d"),TL)
        res = cell(ws,r,6,"Ej testat",YELLOW_BG,bfont(9,True,"374151"),C)
        dv.add(res)
    return ws

# T1–T2: oförändrade från v310
svit("T1 – Importflöde", "T1", "Importflöde & ordningsguide",
     "Testar att filer importeras korrekt, att ordningsguiden visar rätt status och att varningar visas vid fel ordning.",
[
    ("T1.1","Ordningsguide visas på Importera-fliken","1. Gå till Importera-fliken\n2. Titta längst upp","Kortet '📋 Importordning — veckorutin' visas med 4 steg"),
    ("T1.2","Transaktionsfil importeras","1. Tryck Transaktioner\n2. Välj transaktionsfil\n3. Kolla ordningsguiden","Steg 1 visar ✅ och senaste importdatum"),
    ("T1.3","Positionsfil importeras","1. Tryck Positioner\n2. Välj positionsfil","Steg 2 visar ✅. Logg visar tips om Avstämning."),
    ("T1.4","Varning vid inköpskurs utan positioner","1. Rensa appen\n2. Försök importera inköpskursfil UTAN positioner","❌-felmeddelande: 'Inga innehav — importera positionsfil (steg 2) först'"),
    ("T1.5","Inköpskurs efter positioner","1. Importera positioner\n2. Importera inköpskursfil","Steg 3 visar ✅. GAV-historik uppdateras."),
    ("T1.6","Flera positionsfiler bygger historik","1. Importera 3 positionsfiler med olika datum\n2. Gå till Dashboard → Allt","Diagram visar 3+ datapunkter"),
    ("T1.7","Sålda innehav tas bort","1. Importera äldre fil med fler innehav\n2. Importera nyare utan ett innehav","Borttaget innehav loggas: '📤 Borttagna: ...'"),
    ("T1.8","Dubbel import hanteras","1. Importera samma transaktionsfil två gånger","Inga dubbletter — antal ökar inte"),
])

svit("T2 – Dashboard", "T2", "Dashboard & portföljdiagram",
     "Testar det interaktiva diagrammet, nyckeltalskorten och signalbandet.",
[
    ("T2.1","Huvud-diagram visas med default 'i År'","1. Gå till Dashboard","Diagram visas. Aktiv period = 'i År'. Blå portföljlinje."),
    ("T2.2","Periodknappar byter period","1. Tryck '30D'\n2. Tryck 'Allt'","Diagrammet uppdateras. Aktiv knapp markeras."),
    ("T2.3","Kryssrutor lägger till serier","1. Kryssa i 'Nettoinsatt kapital'\n2. Kryssa i 'Nettoresultat'","Nya linjer tillkommer i diagrammet"),
    ("T2.4","Kategorilinjer aktiveras","1. Kryssa i en kategori","Linje i kategorifärgen om historik med katVärden finns"),
    ("T2.5","Stapeldiagram fungerar","1. Tryck '📊 Stapel'","Diagrammet byter till staplar"),
    ("T2.6","Nyckeltalskort","1. Kontrollera 7 kort","💼 📈 💳 📊 🏦 🌱 ✅ med ikon, värde och förklaringsrad"),
    ("T2.7","Global period ändrar nyckeltal","1. Välj '30D'","Avkastningskortet visar '30D'-värde"),
    ("T2.8","Signalband vid röda signaler","1. Innehav med kurs under MA200 i kat. 3–6 och tvådagars aktiv","Rött band '🔴 1 säljsignal' visas"),
    ("T2.9","Tomt diagram hanteras","1. Rensa historik\n2. Gå till Dashboard","'Importera positionsfiler för att se historik'"),
])

# T3: uppdaterad för v311 (inline-inmatning, sparkonto som separat rad)
svit("T3 – Kassa och Avstämning", "T3", "Kassa & Avstämning (v3.11)",
     "Testar Kassa med inline-inmatning per rad, sparkonto som separat rad, och diff-wizarden.",
[
    ("T3.1","Alla konton alltid synliga utan positionsfil",
     "1. Utan att ha importerat positionsfil\n2. Gå till Kassa-fliken",
     "Alla 5 ISK/KF-konton visas med ⚠️-ikon och '—'. Sparkontot visas dimmat längst ned."),
    ("T3.2","Kassa hämtar värden från positionsfil",
     "1. Importera positionsfil\n2. Gå till Kassa",
     "ISK/KF-konton med likvida medel visar 📄-ikon och belopp. Konton utan likvida medel visar ⚠️."),
    ("T3.3","Sparkontot visas som separat rad",
     "1. Gå till Kassa-fliken\n2. Titta i tabellen",
     "'Avanza sparande Martin' visas sist med text '(sparkonto · ej inkl. i totalt)' — dimmat."),
    ("T3.4","Inline-inmatning per rad",
     "1. Hitta ett ISK/KF-konto i tabellen\n2. Skriv ett belopp i inmatningsfältet\n3. Tryck 'Spara'",
     "Kontot visar ✅-ikon och det angivna beloppet. Totalt uppdateras."),
    ("T3.5","Enter-tangent sparar",
     "1. Skriv belopp i inmatningsfältet för ett konto\n2. Tryck Enter",
     "Beloppet sparas (✅-ikon) — som att klicka Spara-knappen."),
    ("T3.6","Alla 6 konton i Avstämning",
     "1. Gå till Avstämning-fliken\n2. Se Saldon per konto",
     "Alla 6 konton visas. Sparkontot visar kontonummer 0040080455 som subrad."),
    ("T3.7","Diff-wizard steg 3 — liten diff",
     "1. Ange siffror från Avanza i steg 2\n2. Tryck 'Beräkna diff →'\n3. Diff < 200 kr",
     "Grönt summakort: '✅ Allt stämmer'."),
    ("T3.8","Manuell override i Kassa sparas",
     "1. Ange belopp inline för ett konto\n2. Importera ny positionsfil",
     "Manuellt värde (✅) skriver över positionsfilens värde (📄) tills manuell override raderas."),
    ("T3.9","Sparkontot kontonummer korrekt",
     "1. Gå till Avstämning\n2. Se kontonummer för Avanza sparande Martin",
     "Kontonummer visas som 0040080455 (med ledande nollor) — inte 40080455."),
])

svit("T4 – Innehav & Signaler", "T4", "Innehav & Signaler",
     "Testar innehavstabellen, MA200-signaler och nödutgångar.",
[
    ("T4.1","Innehavstabellen visas","1. Importera positionsfil\n2. Gå till Innehav","Tabell: Namn, Kategori, Antal, GAV (SEK), Kurs, MA200, Signal, Andel"),
    ("T4.2","Grön MA200-signal","1. Innehav med kurs > 5% över MA200","🟢 visas i signalcell"),
    ("T4.3","Röd signal tvådagarsregel (kat. 3–6)","1. Kurs under MA200\n2. Importera nästa dag","Dag 1: 🟡. Dag 2+: 🔴 i signalbandet"),
    ("T4.4","Nödutgång hård","1. Kat. 4: GAV = 100, kurs = 89 (under 90%)","🛑 Nödutgång hård"),
    ("T4.5","Ticker-sökning","1. Gå till Innehav → Tickers\n2. Klicka 🔍","Lista med börsalternativ visas"),
    ("T4.6","MA200-hämtning","1. Ange API-nyckel\n2. Uppdatera alla MA200","MA200 uppdateras utan valutavarning för SEK"),
    ("T4.7","FX-motor","1. Granska ett USD-innehav","Bolagsvinst och FX-vinst visas separat"),
])

svit("T5 – Excel Backup & Restore", "T5", "Excel-backup & Restore",
     "Testar att all data exporteras korrekt och kan återställas fullständigt.",
[
    ("T5.1","Export skapar fil med 7 ark","1. Backup → 'Ladda ner Excel-backup'","7 ark: Innehav, Transaktioner, Beslutslogg, Sammanfattning, Historik, Kassa, Kategorier"),
    ("T5.2","Historik-arket innehåller totalVärde","1. Öppna Excel\n2. Arket 'Historik'","Kolumn B innehåller faktiska tal. Kolumn D innehåller JSON."),
    ("T5.3","Kassa-arket bevarar kontoStartsaldo","1. Ange manuellt saldo\n2. Exportera","Arket 'Kassa' innehåller 'KontoStartsaldo' med kontot"),
    ("T5.4","Kategorier-arket bevarar anpassningar","1. Redigera en kategori\n2. Exportera","Anpassade värden visas i 'Kategorier'-arket"),
    ("T5.5","Full restore","1. Exportera\n2. Rensa ALLT\n3. Importera backup","Alla innehav, transaktioner, logg och historik återställs"),
    ("T5.6","Diagram efter restore","1. Utför T5.5\n2. Dashboard → Allt","Värdeutvecklingsdiagrammet visar historiklinjen — inte tomt"),
    ("T5.7","Kategoriutveckling efter restore","1. Utför T5.5\n2. Kategorier → Jämför","Kategoriutvecklingsdiagram visar linjer"),
    ("T5.8","Kategorier återställs","1. Utför T5.5\n2. Kategorier → Hantera","Anpassade kategorier med färger/mål återställs"),
    ("T5.9","Restore rensar positionsKassa","1. Utför T5.5\n2. Kassa","Manuella saldo (✅) återställs. Positionsfilsvärden (📄) kräver ny import."),
    ("T5.10","Backup på ny enhet","1. Exportera A\n2. Importera B","Portföljvärde, innehav och historik stämmer"),
])

svit("T6 – Kontoregister", "T6", "Kontoregister & namnbyte",
     "Testar att appen hanterar när ett konto döps om i Avanza.",
[
    ("T6.1","Register byggs vid positionsimport","1. Importera positionsfil\n2. localStorage 'portfölj_konto_register'","JSON med kontonummer som nycklar"),
    ("T6.2","Korrekt namn i Avstämning","1. Importera positionsfil\n2. Avstämning → Saldon","Kontonamnen matchar Avanza"),
    ("T6.3","Namnbyte detekteras","1. Ändra kontonamn i CSV\n2. Importera","Register uppdateras med nytt namn och historik"),
    ("T6.4","Historikindikator visas","1. Utför T6.3\n2. Avstämning","🔁 1 tidigare namn med tooltip"),
    ("T6.5","avanzaKontoInfo via register","1. Utför T6.3\n2. Kassa och Innehav","Konto identifieras korrekt via register"),
])

svit("T7 – Mobil iPad", "T7", "Mobilanvändning på iPad",
     "Testar att alla funktioner fungerar på iPad i Safari.",
[
    ("T7.1","Filimport i Safari","1. Öppna i Safari på iPad\n2. Tryck Positioner","Filväljar-dialog öppnas i iOS Filer-appen"),
    ("T7.2","Flervalsimport","1. Välj flera filer","Alla valda filer importeras"),
    ("T7.3","Dashboard-diagrammet touchvänligt","1. Tryck på diagrammet på iPad","Tooltip visas. Responsivt."),
    ("T7.4","Period-knappar på iPad","1. Tryck period-knapp","Knappen aktiveras ≥ 44px touch-yta"),
    ("T7.5","Mobilkort i Innehav","1. Sök ett innehav på iPad","Mobilkort visas med alla fält"),
    ("T7.6","Bottom-navigering","1. Tryck flikarna","Rätt sektion visas. Aktiv flik markeras."),
])

svit("T8 – Veckorutin", "T8", "Veckorutin — komplett helflöde",
     "Simulerar hela veckorutinen från Avanza till loggpost.",
[
    ("T8.1","Steg 1: Exportera från Avanza","1. avanza.se → Transaktioner → Exportera\n2. Innehav → Exportera","Två CSV-filer sparas"),
    ("T8.2","Steg 2: Importera transaktioner","1. Importera → Transaktioner","✅ steg 1. Rätt antal transaktioner."),
    ("T8.3","Steg 2: Importera positioner","1. Importera → Positioner","✅ steg 2. Kassa uppdateras. Historikpunkt tillagd."),
    ("T8.4","Steg 3: Stäm av","1. Avstämning → diff-kontroll\n2. Ange Avanza-totaler","Diff < 1%."),
    ("T8.5","Kontrollera signalband","1. Dashboard → signalbandet","Inga röda ELLER röd signal identifierad"),
    ("T8.6","Kontrollera kategoribalans","1. Kategorier → viktbalans","Ingen ombalanseringsvarning ELLER tydlig indikation"),
    ("T8.7","Skriv beslutslogg","1. Beslutslogg → fyll i\n2. Spara","Loggpost sparas och visas"),
])

svit("T9 – Felhantering", "T9", "Felhantering & felmeddelanden",
     "Testar att appen hanterar fel gracefully.",
[
    ("T9.1","Fel filtyp","1. Välj transaktionsfil vid positionsimport","❌ 'Det här verkar inte vara en positionsfil'"),
    ("T9.2","Inköpskurs utan positioner","1. Rensa\n2. Importera inköpskurs","❌ 'Importera positionsfil (steg 2) först'"),
    ("T9.3","Okänt konto filtreras","1. Lägg till okänt kontonummer i CSV\n2. Importera","Loggen: 'X exkluderade av filter'"),
    ("T9.4","Fel Excel-format","1. Välj icke-backup Excel","Felmeddelande om saknade ark"),
    ("T9.5","Tomt portfölj","1. Rensa all data\n2. Gå igenom alla flikar","Tomma tillstånd med förklarande text. Inga JS-fel."),
    ("T9.6","API-fel vid MA200","1. Ogiltig API-nyckel\n2. Uppdatera alla MA200","Felmeddelande per innehav. Appen kraschar inte."),
])

# T10: NYA testfall för Inställningar-sektionen
svit("T10 – Inställningar", "T10", "Inställningar-sektionen (ny i v3.11)",
     "Testar alla delar av den nya Inställningar-sektionen: kontokonfig, kategori-editor, strategiparametrar, profil, VP-filter och export/import.",
[
    ("T10.1","Inställningar nås via nav",
     "1. Tryck på ⚙️ Inställningar i navigeringsmenyn",
     "Sektionen öppnas med 6 kort: Kontokonfiguration, Kategori-editor, Strategiparametrar, Profil, Värdepappersfilter, Export/import"),
    ("T10.2","Kontokonfiguration: redigera konto",
     "1. Inställningar → Kontokonfiguration\n2. Tryck ✏️ på ett konto\n3. Ändra namn\n4. Tryck ✓ Spara",
     "Kontot uppdateras i listan. Ändringen syns direkt i Kassa och Avstämning."),
    ("T10.3","Kontokonfiguration: lägg till konto",
     "1. Tryck '+ Lägg till konto'\n2. Fyll i namn, kontonummer, typ ISK\n3. Spara",
     "Nytt konto läggs till i listan. Syns i Kassa-tabellen vid nästa import."),
    ("T10.4","Kontokonfiguration: återställ fabriksinst.",
     "1. Gör en ändring\n2. Tryck '↺ Återställ fabriksinställningar'\n3. Bekräfta",
     "De 6 originalkontona återställs. Portföljdata orörd."),
    ("T10.5","Kategori-editor: redigera kategori",
     "1. Inställningar → Kategori-editor\n2. Tryck ✏️ på en kategori\n3. Ändra namn och emoji\n4. Spara",
     "Kategorin uppdateras. Ändringen syns direkt i Dashboard, Innehav och Signaler."),
    ("T10.6","Kategori-editor: lägg till kategori",
     "1. Tryck '+ Lägg till kategori'\n2. Fyll i alla fält\n3. Spara",
     "Ny kategori syns i listan och i Kategorier-fliken. Kan tilldelas innehav."),
    ("T10.7","Strategiparametrar: ändra MA200-gräns",
     "1. Inställningar → Strategiparametrar\n2. Ändra 'MA200 — grön signal' från 5 till 3",
     "Fältet markeras med blå kant. Signalstatus för innehav uppdateras direkt."),
    ("T10.8","Strategiparametrar: återställ enstaka",
     "1. Ändra ett värde\n2. Tryck '↺ 5'-knappen (återställ)",
     "Enbart det fältet återgår till standardvärdet. Övriga behålls."),
    ("T10.9","Export av inställningar",
     "1. Inställningar → Export/import\n2. Tryck 'Exportera inställningar (JSON)'",
     "JSON-fil laddas ned: 'installningar_strategiportfoljen_DATUM.json'. Innehåller konton, kategorier, params, profil, filter."),
    ("T10.10","Import av inställningar",
     "1. Exportera inställningar (T10.9)\n2. Gör ändringar i appen\n3. Importera JSON-filen",
     "Inställningarna återställs till exportfilens värden. Portföljdata (innehav, historik) orörd."),
])

dst = r"C:\Users\hejma\Projekt_Claude\strategiportfoljen\Testprotokoll_Strategiportfoljen_v311.xlsx"
wb.save(dst)
print(f"Skapad: {dst}")
print(f"Totalt: {sum(n for _,n in sviter)} testfall i 10 sviter")
