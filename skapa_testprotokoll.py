"""
Skapar Testprotokoll_Strategiportfoljen_v208.xlsx
"""
from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# ── Färger ──────────────────────────────────────────────────
NAVY      = "1E3A5F"
ACCENT    = "0EA5E9"
HEADER_BG = "1E3A5F"
SUITE_BG  = "2563EB"
ROW_ALT   = "F0F7FF"
ROW_WHITE = "FFFFFF"
GREEN_BG  = "DCFCE7"
RED_BG    = "FEE2E2"
YELLOW_BG = "FEF9C3"
GRAY_BG   = "F3F4F6"
BORDER_C  = "E5E7EB"

def col_fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def thin_border(top=True, bottom=True, left=True, right=True):
    s = Side(style="thin", color=BORDER_C)
    return Border(
        top=s if top else None,
        bottom=s if bottom else None,
        left=s if left else None,
        right=s if right else None,
    )

def header_font(size=11, bold=True, color="FFFFFF"):
    return Font(name="Calibri", size=size, bold=bold, color=color)

def body_font(size=10, bold=False, color="111827"):
    return Font(name="Calibri", size=size, bold=bold, color=color)

def apply_cell(ws, row, col, value="", fill=None, font=None,
               align=None, border=None, wrap=False):
    c = ws.cell(row=row, column=col, value=value)
    if fill:   c.fill   = fill
    if font:   c.font   = font
    if border: c.border = border
    if align:  c.alignment = align
    elif wrap: c.alignment = Alignment(wrap_text=True, vertical="top")
    return c

CENTER = Alignment(horizontal="center", vertical="center")
LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)
TOP_L  = Alignment(horizontal="left",   vertical="top",    wrap_text=True)

wb = Workbook()

# ════════════════════════════════════════════════════════════
#  HJÄLPFUNKTION: lägg till en testsvit-flik
# ════════════════════════════════════════════════════════════
def skapa_svit(wb, sheet_title, suite_code, suite_title, suite_desc,
               forutsattning, steg_lista, kolumner, rader, notering_rader=3):
    """
    kolumner: lista av (rubrik, bredd, fillable?)
    rader:    lista av tupler med cellvärden (en per kolumn)
    """
    ws = wb.create_sheet(title=sheet_title)
    ws.sheet_view.showGridLines = False

    # ── Sidmarginaler ──
    ws.column_dimensions["A"].width = 5
    for i, (_, w, _) in enumerate(kolumner, start=2):
        ws.column_dimensions[get_column_letter(i)].width = w

    r = 1

    # ── Stor rubrik ──
    ws.merge_cells(start_row=r, start_column=1, end_row=r,
                   end_column=1 + len(kolumner))
    c = ws.cell(row=r, column=1,
                value=f"{suite_code} — {suite_title}")
    c.fill  = col_fill(NAVY)
    c.font  = Font(name="Calibri", size=16, bold=True, color="FFFFFF")
    c.alignment = Alignment(horizontal="left", vertical="center",
                            indent=1)
    ws.row_dimensions[r].height = 36
    r += 1

    # ── Beskrivning ──
    ws.merge_cells(start_row=r, start_column=1, end_row=r,
                   end_column=1 + len(kolumner))
    c = ws.cell(row=r, column=1, value=suite_desc)
    c.fill = col_fill("EFF6FF")
    c.font = Font(name="Calibri", size=10, italic=True, color="1E3A5F")
    c.alignment = Alignment(horizontal="left", vertical="center",
                            indent=1, wrap_text=True)
    ws.row_dimensions[r].height = 28
    r += 1

    # ── Förutsättning ──
    if forutsattning:
        ws.merge_cells(start_row=r, start_column=1, end_row=r,
                       end_column=1 + len(kolumner))
        c = ws.cell(row=r, column=1,
                    value="⚠️  Förutsättning: " + forutsattning)
        c.fill = col_fill("FEF3C7")
        c.font = Font(name="Calibri", size=9, color="92400E")
        c.alignment = Alignment(horizontal="left", vertical="center",
                                indent=1, wrap_text=True)
        ws.row_dimensions[r].height = 22
        r += 1

    # ── Steg ──
    ws.merge_cells(start_row=r, start_column=1, end_row=r,
                   end_column=1 + len(kolumner))
    c = ws.cell(row=r, column=1, value="TESTSTEG")
    c.fill = col_fill(SUITE_BG)
    c.font = Font(name="Calibri", size=9, bold=True, color="FFFFFF")
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[r].height = 18
    r += 1

    for i, steg in enumerate(steg_lista, 1):
        ws.merge_cells(start_row=r, start_column=1, end_row=r,
                       end_column=1 + len(kolumner))
        c = ws.cell(row=r, column=1, value=f"  {i}.  {steg}")
        c.fill = col_fill(ROW_WHITE if i % 2 else ROW_ALT)
        c.font = Font(name="Calibri", size=9, color="374151")
        c.alignment = Alignment(horizontal="left", vertical="top",
                                indent=1, wrap_text=True)
        ws.row_dimensions[r].height = 20
        r += 1

    r += 1  # blank

    # ── Kolumnrubriker ──
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=1)
    for ci, (rubrik, _, _) in enumerate(kolumner, start=2):
        c = ws.cell(row=r, column=ci, value=rubrik)
        c.fill = col_fill(NAVY)
        c.font = Font(name="Calibri", size=9, bold=True, color="FFFFFF")
        c.alignment = CENTER
        c.border = thin_border()
    # Nr-kolumn
    c = ws.cell(row=r, column=1, value="#")
    c.fill = col_fill(NAVY)
    c.font = Font(name="Calibri", size=9, bold=True, color="FFFFFF")
    c.alignment = CENTER
    c.border = thin_border()
    ws.row_dimensions[r].height = 22
    hdr_row = r
    r += 1

    # ── Datarader ──
    ok_col = 1 + len(kolumner)   # sista kolumnen = OK?
    for ri, rad in enumerate(rader):
        bg = ROW_WHITE if ri % 2 == 0 else ROW_ALT
        # Nr
        c = ws.cell(row=r, column=1, value=rad[0])
        c.fill = col_fill(bg)
        c.font = Font(name="Calibri", size=9, bold=True, color=NAVY)
        c.alignment = CENTER
        c.border = thin_border()
        # Övriga kolumner
        for ci, (val, (_, _, fillable)) in enumerate(
                zip(rad[1:], kolumner), start=2):
            c = ws.cell(row=r, column=ci, value=val)
            if ci == ok_col:
                c.fill = col_fill("F0FDF4")
                c.font = Font(name="Calibri", size=11, color="16A34A")
                c.alignment = CENTER
            else:
                c.fill = col_fill(GRAY_BG if fillable else bg)
                c.font = Font(name="Calibri", size=9,
                              color="374151" if not fillable else "111827")
                c.alignment = TOP_L
            c.border = thin_border()
        ws.row_dimensions[r].height = 30
        r += 1

    r += 1  # blank

    # ── Noteringar ──
    ws.merge_cells(start_row=r, start_column=1, end_row=r,
                   end_column=1 + len(kolumner))
    c = ws.cell(row=r, column=1, value="Noteringar / avvikelser:")
    c.fill = col_fill(SUITE_BG)
    c.font = Font(name="Calibri", size=9, bold=True, color="FFFFFF")
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[r].height = 18
    r += 1

    for _ in range(notering_rader):
        ws.merge_cells(start_row=r, start_column=1, end_row=r,
                       end_column=1 + len(kolumner))
        c = ws.cell(row=r, column=1, value="")
        c.fill = col_fill(ROW_WHITE)
        c.border = thin_border()
        ws.row_dimensions[r].height = 22
        r += 1

    # ── Frys rubrikrad ──
    ws.freeze_panes = ws.cell(row=hdr_row + 1, column=1)

    return ws


# ════════════════════════════════════════════════════════════
#  ARK 0 — FRAMSIDA
# ════════════════════════════════════════════════════════════
ws0 = wb.active
ws0.title = "Framsida"
ws0.sheet_view.showGridLines = False
ws0.column_dimensions["A"].width = 2
ws0.column_dimensions["B"].width = 28
ws0.column_dimensions["C"].width = 32
ws0.column_dimensions["D"].width = 20
ws0.column_dimensions["E"].width = 20

r = 1
ws0.merge_cells("B1:E3")
c = ws0["B1"]
c.value = "Testprotokoll"
c.fill  = col_fill(NAVY)
c.font  = Font(name="Calibri", size=28, bold=True, color="FFFFFF")
c.alignment = Alignment(horizontal="center", vertical="center")
ws0.row_dimensions[1].height = 30
ws0.row_dimensions[2].height = 30
ws0.row_dimensions[3].height = 30

ws0.merge_cells("B4:E4")
c = ws0["B4"]
c.value = "Strategiportföljen v2.08 — Excel Backup & Återställning"
c.fill  = col_fill(ACCENT)
c.font  = Font(name="Calibri", size=13, bold=True, color="FFFFFF")
c.alignment = Alignment(horizontal="center", vertical="center")
ws0.row_dimensions[4].height = 28

r = 6
for label, val_placeholder, fill_c in [
    ("Testare",         "", True),
    ("Testdatum",       "", True),
    ("Appversion",      "v2.08", False),
    ("Miljö (t.ex. Safari iPad / Chrome dator)", "", True),
]:
    ws0.row_dimensions[r].height = 26
    c = ws0.cell(row=r, column=2, value=label)
    c.fill  = col_fill(GRAY_BG)
    c.font  = Font(name="Calibri", size=10, bold=True, color=NAVY)
    c.alignment = LEFT
    c.border = thin_border()

    c = ws0.cell(row=r, column=3, value=val_placeholder)
    c.fill  = col_fill(ROW_WHITE if fill_c else GRAY_BG)
    c.font  = Font(name="Calibri", size=10, color="374151")
    c.alignment = LEFT
    c.border = thin_border()
    ws0.merge_cells(start_row=r, start_column=3, end_row=r, end_column=5)
    r += 1

r += 1
ws0.merge_cells(start_row=r, start_column=2, end_row=r, end_column=5)
c = ws0.cell(row=r, column=2, value="TESTSVITER — SAMMANFATTNING")
c.fill  = col_fill(SUITE_BG)
c.font  = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
ws0.row_dimensions[r].height = 20
r += 1

sum_hdr = ["Testsvit", "Antal testfall", "Godkända", "Underkända"]
for ci, h in enumerate(sum_hdr, start=2):
    c = ws0.cell(row=r, column=ci, value=h)
    c.fill  = col_fill(NAVY)
    c.font  = Font(name="Calibri", size=9, bold=True, color="FFFFFF")
    c.alignment = CENTER
    c.border = thin_border()
ws0.row_dimensions[r].height = 20
r += 1

sviter = [
    ("T1 — Export: grundläggande", 7),
    ("T2 — Export: dataintegritet", 9),
    ("T3 — Export: header-knapp",  2),
    ("T4 — Rundtur: export→rensa→import", 10),
    ("T5 — Felhantering", 6),
    ("T6 — Äldre exportformat", 2),
    ("T7 — Mobil / iPad", 6),
    ("T8 — Kategorier bevaras", 4),
]
totalt = sum(n for _, n in sviter)

for i, (namn, antal) in enumerate(sviter):
    bg = ROW_WHITE if i % 2 == 0 else ROW_ALT
    ws0.row_dimensions[r].height = 22
    for ci, val in enumerate([namn, antal, "", ""], start=2):
        c = ws0.cell(row=r, column=ci, value=val)
        c.fill  = col_fill(GRAY_BG if ci in (4, 5) else bg)
        c.font  = Font(name="Calibri", size=9, color="374151")
        c.alignment = CENTER if ci > 2 else LEFT
        c.border = thin_border()
    r += 1

# Totalrad
ws0.row_dimensions[r].height = 22
for ci, val in enumerate(["TOTALT", totalt, "", ""], start=2):
    c = ws0.cell(row=r, column=ci, value=val)
    c.fill  = col_fill(NAVY)
    c.font  = Font(name="Calibri", size=9, bold=True, color="FFFFFF")
    c.alignment = CENTER if ci > 2 else LEFT
    c.border = thin_border()
r += 2

# Godkänd-rad
ws0.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
c = ws0.cell(row=r, column=2, value="Godkänd? (Ja / Nej / Med anmärkningar)")
c.fill  = col_fill(GRAY_BG)
c.font  = Font(name="Calibri", size=10, bold=True, color=NAVY)
c.alignment = LEFT
c.border = thin_border()
ws0.merge_cells(start_row=r, start_column=4, end_row=r, end_column=5)
c = ws0.cell(row=r, column=4, value="")
c.fill  = col_fill(ROW_WHITE)
c.border = thin_border()
ws0.row_dimensions[r].height = 26
r += 1

ws0.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
c = ws0.cell(row=r, column=2, value="Signatur")
c.fill  = col_fill(GRAY_BG)
c.font  = Font(name="Calibri", size=10, bold=True, color=NAVY)
c.alignment = LEFT
c.border = thin_border()
ws0.merge_cells(start_row=r, start_column=4, end_row=r, end_column=5)
c = ws0.cell(row=r, column=4, value="")
c.fill  = col_fill(ROW_WHITE)
c.border = thin_border()
ws0.row_dimensions[r].height = 26


# ════════════════════════════════════════════════════════════
#  T4 — NYCKELTALSJÄMFÖRELSE (eget ark)
# ════════════════════════════════════════════════════════════
def skapa_nyckeltal_ark(wb):
    ws = wb.create_sheet("T4 – Nyckeltal")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 38
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 22
    ws.column_dimensions["E"].width = 12

    r = 1
    ws.merge_cells("A1:E2")
    c = ws["A1"]
    c.value = "T4 — Nyckeltalsjämförelse (fyll i INNAN rensning)"
    c.fill  = col_fill(NAVY)
    c.font  = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 24
    ws.row_dimensions[2].height = 24
    r = 3

    ws.merge_cells(f"A{r}:E{r}")
    c = ws.cell(row=r, column=1,
        value="Notera värdena i appen INNAN du rensar. Fyll i EFTER-kolumnen efter återställning.")
    c.fill  = col_fill("FEF3C7")
    c.font  = Font(name="Calibri", size=9, italic=True, color="92400E")
    c.alignment = Alignment(horizontal="left", vertical="center",
                            indent=1, wrap_text=True)
    ws.row_dimensions[r].height = 22
    r += 2

    hdrs = ["", "Nyckeltal", "Värde FÖRE rensning", "Värde EFTER import", "Matchar?"]
    for ci, h in enumerate(hdrs, start=1):
        c = ws.cell(row=r, column=ci, value=h)
        c.fill  = col_fill(NAVY)
        c.font  = Font(name="Calibri", size=9, bold=True, color="FFFFFF")
        c.alignment = CENTER
        c.border = thin_border()
    ws.row_dimensions[r].height = 22
    r += 1

    nyckeltal = [
        "Antal innehav",
        "Totalt portföljvärde",
        "Innehavens värde",
        "Nettoinsatt kapital",
        "Tillgänglig likviditet",
        "Antal transaktioner (Transaktioner-fliken)",
        "Antal loggposter (Beslutslogg-fliken)",
        "Antal historikpunkter (Importera → Historikposter)",
    ]
    for i, nk in enumerate(nyckeltal):
        bg = ROW_WHITE if i % 2 == 0 else ROW_ALT
        ws.row_dimensions[r].height = 26
        c = ws.cell(row=r, column=1, value=f"{i+1}")
        c.fill  = col_fill(bg)
        c.font  = Font(name="Calibri", size=9, bold=True, color=NAVY)
        c.alignment = CENTER
        c.border = thin_border()

        c = ws.cell(row=r, column=2, value=nk)
        c.fill  = col_fill(bg)
        c.font  = Font(name="Calibri", size=9, color="374151")
        c.alignment = LEFT
        c.border = thin_border()

        for ci in [3, 4]:
            c = ws.cell(row=r, column=ci, value="")
            c.fill  = col_fill(GRAY_BG)
            c.font  = Font(name="Calibri", size=10, color="111827")
            c.alignment = CENTER
            c.border = thin_border()

        c = ws.cell(row=r, column=5, value="☐")
        c.fill  = col_fill("F0FDF4")
        c.font  = Font(name="Calibri", size=12, color="16A34A")
        c.alignment = CENTER
        c.border = thin_border()
        r += 1

skapa_nyckeltal_ark(wb)


# ════════════════════════════════════════════════════════════
#  TESTSVITER T1–T8
# ════════════════════════════════════════════════════════════

# Standardkolumner för de flesta sviter
STD_KOLS = [
    ("Vad kontrolleras",  38, False),
    ("Förväntat resultat", 36, False),
    ("Faktiskt utfall",    28, True),
    ("OK? (☑/✗)",         10, True),
]

# ── T1 ──────────────────────────────────────────────────────
skapa_svit(wb,
    sheet_title="T1 – Export grundläggande",
    suite_code="T1",
    suite_title="Export: grundläggande funktionalitet",
    suite_desc="Kontrollerar att exportknappen skapar en giltig Excel-fil med rätt antal ark och korrekt metadata.",
    forutsattning="Det finns befintliga innehav och transaktioner i portföljen.",
    steg_lista=[
        "Gå till Importera-fliken → Backup & Återställning",
        "Tryck '⬇ Ladda ner Excel-backup'",
        "Öppna den nedladdade filen portfölj_DATUM.xlsx i Excel eller Numbers",
        "Räkna antal ark och kontrollera deras namn",
        "Öppna arket Sammanfattning och granska rad 1–2",
        "Öppna arket Innehav och granska kolumnrubrikerna",
    ],
    kolumner=STD_KOLS,
    rader=[
        ("1.1", "Fil laddas ned automatiskt", "Fil med namn portfölj_ÅÅÅÅ-MM-DD.xlsx", "", "☐"),
        ("1.2", "Antal ark i filen", "7 ark", "", "☐"),
        ("1.3", "Arknamn", "Innehav · Transaktioner · Beslutslogg · Sammanfattning · Historik · Kassa · Kategorier", "", "☐"),
        ("1.4", "Sammanfattning rad 1", "BACKUP_VERSION = 2.08", "", "☐"),
        ("1.5", "Sammanfattning rad 2", "BACKUP_DATUM = ISO-datumstämpel", "", "☐"),
        ("1.6", "Innehav: obligatoriska kolumner", "id, Namn, Antal, GAV_SEK, Kurs_SEK, Kategori, GAV_Lokal, Historical_FX, TvåDagarsAktiv", "", "☐"),
        ("1.7", "Importlogg längst ned på sidan", "Grön rad: 'Excel-backup exporterad: X innehav, Y transaktioner'", "", "☐"),
    ],
)

# ── T2 ──────────────────────────────────────────────────────
skapa_svit(wb,
    sheet_title="T2 – Export dataintegritet",
    suite_code="T2",
    suite_title="Export: dataintegritet per ark",
    suite_desc="Kontrollerar att varje ark innehåller rätt antal rader och att värdena stämmer med vad appen visar.",
    forutsattning="Notera antal innehav, transaktioner och loggposter i appen INNAN export.",
    steg_lista=[
        "Notera antal innehav i Innehav-fliken",
        "Notera antal transaktioner i Transaktioner-fliken",
        "Notera antal loggposter i Beslutslogg-fliken",
        "Exportera backup",
        "Jämför radantal i varje ark med appens värden",
        "Verifiera ett slumpmässigt innehav (GAV_SEK stämmer med tabellen)",
        "Jämför Sammanfattning → Innehavens marknadsvärde mot Dashboard",
    ],
    kolumner=[
        ("Vad kontrolleras",  30, False),
        ("Förväntat",         22, False),
        ("Appens värde",      18, True),
        ("Excel-värde",       18, True),
        ("OK?",               10, True),
    ],
    rader=[
        ("2.1", "Innehav: antal datarader",        "= antal innehav i appen",        "", "", "☐"),
        ("2.2", "Transaktioner: antal datarader",   "= antal transaktioner i appen",  "", "", "☐"),
        ("2.3", "Beslutslogg: antal datarader",     "= antal loggposter i appen",     "", "", "☐"),
        ("2.4", "Historik: antal datarader",        "= antal historikpunkter",         "", "", "☐"),
        ("2.5", "Kassa ark rad 1",                  "SEKTION / KassaTransaktioner",   "", "", "☐"),
        ("2.6", "Kassa ark: andra sektionsrubrik",  "SEKTION / KontoStartsaldo",      "", "", "☐"),
        ("2.7", "Kategorier: antal rader",          "= antal kategorier (standard: 6)","", "", "☐"),
        ("2.8", "Slumpmässigt innehav — GAV_SEK",   "Matchar värde i Innehav-tabellen","", "", "☐"),
        ("2.9", "Sammanfattning: marknadsvärde",    "Matchar 'Innehavens värde' på Dashboard","", "", "☐"),
    ],
)

# ── T3 ──────────────────────────────────────────────────────
skapa_svit(wb,
    sheet_title="T3 – Header-knapp",
    suite_code="T3",
    suite_title="Export: header-knappen ⬇ Excel",
    suite_desc="Kontrollerar att exportknappen uppe till höger i headern fungerar identiskt med knappen i Importera-sektionen.",
    forutsattning=None,
    steg_lista=[
        "Tryck knappen '⬇ Excel' i appens header (övre högra hörnet)",
        "Öppna den nedladdade filen och kontrollera arkstruktur",
    ],
    kolumner=STD_KOLS,
    rader=[
        ("3.1", "Fil skapas",         "Fil laddas ned med korrekt namn",  "", "☐"),
        ("3.2", "7 ark med rätt namn", "Identisk struktur som T1",        "", "☐"),
    ],
)

# ── T4 ──────────────────────────────────────────────────────
skapa_svit(wb,
    sheet_title="T4 – Rundtur",
    suite_code="T4",
    suite_title="Rundtur: export → rensa → import",
    suite_desc="Huvudtest. Exportera all data, rensa portföljen och importera tillbaka. Verifiera att allt återställts exakt. Fyll i nyckeltal i arket 'T4 – Nyckeltal' INNAN rensning.",
    forutsattning="Fyll i arket 'T4 – Nyckeltal' med appens aktuella värden INNAN du påbörjar testet.",
    steg_lista=[
        "Fyll i fliken 'T4 – Nyckeltal' med appens aktuella värden",
        "Exportera backup via Importera → Backup & Återställning → ⬇ Ladda ner Excel-backup",
        "Gå till Underhåll → 🗑 Rensa all data — bekräfta i dialogen",
        "Verifiera att appen är tom (Dashboard visar '—')",
        "Gå till Backup & Återställning → ⬆ Importera från backup",
        "Välj den exporterade filen — läs bekräftelsedialogens text",
        "Tryck OK",
        "Jämför alla nyckeltal i 'T4 – Nyckeltal'-arket",
    ],
    kolumner=STD_KOLS,
    rader=[
        ("4.1",  "Bekräftelsedialog visas",              "Visar backup-version och datum",                        "", "☐"),
        ("4.2",  "Import genomförs utan felmeddelande",   "Alert: 'Backup återställd!' med statistik",             "", "☐"),
        ("4.3",  "Importlogg",                           "Grön rad med antal återställda poster",                 "", "☐"),
        ("4.4",  "Dashboard: nyckeltal matchar",          "Se T4 – Nyckeltal-arket",                              "", "☐"),
        ("4.5",  "Innehav-fliken: samma antal aktier",    "Se T4 – Nyckeltal-arket",                              "", "☐"),
        ("4.6",  "Transaktioner: samma antal",            "Se T4 – Nyckeltal-arket",                              "", "☐"),
        ("4.7",  "Beslutslogg: samma loggposter",         "Se T4 – Nyckeltal-arket",                              "", "☐"),
        ("4.8",  "Värdeutvecklingsdiagram visas",         "Diagrammet fylls med historikpunkter",                  "", "☐"),
        ("4.9",  "Kategorier: namn och färger",           "Inga kategorier har återgått till standardnamn",        "", "☐"),
        ("4.10", "Kassa-fliken: kontosaldon",             "Manuella kontosaldon matchar vad de var före rensning", "", "☐"),
    ],
)

# ── T5 ──────────────────────────────────────────────────────
skapa_svit(wb,
    sheet_title="T5 – Felhantering",
    suite_code="T5",
    suite_title="Import: validering av felaktiga filer",
    suite_desc="Kontrollerar att appen hanterar felaktiga filer gracefully — inga krascher, tydliga felmeddelanden.",
    forutsattning=None,
    steg_lista=[
        "T5a — Välj en Avanza CSV-fil i backup-importknappen",
        "T5b — Exportera backup, öppna i Excel, ta bort kolumnen 'Namn' i Innehav-arket, spara och importera",
        "T5c — Välj en giltig backup-fil men tryck 'Avbryt' i bekräftelsedialogens",
        "T5d — Exportera backup, öppna i Excel, ta bort alla datarader i Innehav-arket (behåll rubrikraden), importera",
    ],
    kolumner=STD_KOLS,
    rader=[
        ("5a-1", "CSV-fil importeras (fel filtyp)",      "Felmeddelande: arket 'Innehav' saknas",                "", "☐"),
        ("5a-2", "Appen kraschar inte",                  "Fortsätter fungera normalt efter felmeddelandet",      "", "☐"),
        ("5b-1", "Saknad kolumn 'Namn' importeras",      "Felmeddelande listar saknade obligatoriska kolumner",  "", "☐"),
        ("5b-2", "Befintlig data bevaras",               "Portföljdata oförändrad efter misslyckat försök",      "", "☐"),
        ("5c-1", "Avbryt i bekräftelsedialogen",         "Ingen data ändras, portföljen intakt",                 "", "☐"),
        ("5d-1", "Tom Innehav-tabell",                   "Varningsdialogruta visas med fråga om att fortsätta",  "", "☐"),
    ],
)

# ── T6 ──────────────────────────────────────────────────────
skapa_svit(wb,
    sheet_title="T6 – Äldre format",
    suite_code="T6",
    suite_title="Import: äldre exportformat (saknar BACKUP_VERSION)",
    suite_desc="Kontrollerar att en export utan BACKUP_VERSION (t.ex. v2.07) hanteras med varning men ändå fungerar om Innehav-kolumnerna stämmer.",
    forutsattning="Kräver en exportfil från v2.07, ELLER: öppna en ny backup i Excel och ta bort rad 1–2 i Sammanfattning-arket.",
    steg_lista=[
        "Öppna en backup i Excel → Sammanfattning-arket → ta bort raderna med BACKUP_VERSION och BACKUP_DATUM",
        "Spara filen",
        "Importera den modifierade filen via Importera → Backup & Återställning",
        "Observera dialogen och bekräfta",
    ],
    kolumner=STD_KOLS,
    rader=[
        ("6.1", "Dialog om saknad backup-markering visas", "Varningstext med fråga om att fortsätta trots saknad markering", "", "☐"),
        ("6.2", "Import lyckas om Innehav-kolumner finns", "Data återställs korrekt trots saknad version-info",              "", "☐"),
    ],
)

# ── T7 ──────────────────────────────────────────────────────
skapa_svit(wb,
    sheet_title="T7 – Mobil iPad",
    suite_code="T7",
    suite_title="Mobil / iPad",
    suite_desc="Kontrollerar att export och import fungerar i Safari på iPad/iPhone, inklusive interaktion med Filer-appen.",
    forutsattning="Öppna appen i Safari på iPad eller iPhone.",
    steg_lista=[
        "Öppna appen i Safari på iPad",
        "Navigera till Importera via bottom-nav (⬆-ikonen)",
        "Scrolla ned till sektionen 'Backup & Återställning'",
        "Tryck '⬇ Ladda ner Excel-backup' — filen ska sparas i Filer-appen",
        "Tryck '⬆ Importera från backup' — välj filen från Filer-appen",
        "Genomför import och verifiera resultatet",
    ],
    kolumner=STD_KOLS,
    rader=[
        ("7.1", "Backup-sektionen visas",             "Synlig under Importera-fliken",                     "", "☐"),
        ("7.2", "Exportknappen är klickbar",           "Minst 48px hög, tryckbar utan att missa",           "", "☐"),
        ("7.3", "Export: fil hamnar i Filer-appen",    "iOS sparar portfölj_DATUM.xlsx korrekt",            "", "☐"),
        ("7.4", "Import: Filer-appen öppnas vid tryck","Kan navigera till och välja backup-filen",          "", "☐"),
        ("7.5", "Import: bekräftelsedialog visas",     "Samma dialog som på dator",                         "", "☐"),
        ("7.6", "Import: data återställs korrekt",     "Identisk med T4 — alla nyckeltal matchar",          "", "☐"),
    ],
)

# ── T8 ──────────────────────────────────────────────────────
skapa_svit(wb,
    sheet_title="T8 – Kategorier",
    suite_code="T8",
    suite_title="Kategorier bevaras vid backup & återställning",
    suite_desc="Kontrollerar att anpassade kategorier (ej standard) exporteras och importeras korrekt med alla inställningar bevarade.",
    forutsattning="Minst en kategori ska ha ett anpassat namn, färg eller MA200-inställning. Notera värdena i tabellen nedan.",
    steg_lista=[
        "Öppna Kategorier-fliken → välj en anpassad kategori → tryck ✏️ redigera",
        "Notera: namn, färg (hex), MA200-regel och målvikter (Min/Max %)",
        "Exportera backup",
        "Rensa all data",
        "Importera backup",
        "Öppna Kategorier-fliken och jämför med dina noteringar",
    ],
    kolumner=[
        ("Vad kontrolleras",  28, False),
        ("Förväntat",         24, False),
        ("Värde FÖRE",        20, True),
        ("Värde EFTER",       20, True),
        ("OK?",               10, True),
    ],
    rader=[
        ("8.1", "Kategorinamn bevaras",   "Anpassat namn återställt",          "", "", "☐"),
        ("8.2", "Kategori-färg bevaras",  "Samma hex-färg (t.ex. #7C3AED)",    "", "", "☐"),
        ("8.3", "MA200-regel bevaras",    "'ma200' eller 'ingen' korrekt",     "", "", "☐"),
        ("8.4", "Målvikter bevaras",      "Min% och Max% samma som före",      "", "", "☐"),
    ],
)


# ════════════════════════════════════════════════════════════
#  Spara
# ════════════════════════════════════════════════════════════
utfil = r"C:\Users\hejma\Projekt_Claude\strategiportfoljen\Testprotokoll_Strategiportfoljen_v208.xlsx"
wb.save(utfil)
print(f"Skapad: {utfil}")
